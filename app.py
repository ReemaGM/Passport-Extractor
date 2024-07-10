from flask import Flask, request, send_file, render_template, jsonify, Response, url_for
import os
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from mindee import Client, product
from country_codes import country_codes
from mrz.checker.td3 import TD3CodeChecker, get_country
import time

UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'gif', 'xlsx'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(PROCESSED_FOLDER):
    os.makedirs(PROCESSED_FOLDER)

# Initialize Mindee client
mindee_client = Client(api_key="bbb9eca0718ca96c4a99060009b449e6")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

progress = 0

@app.route('/')
def upload_form():
    return render_template('upload_form.html')

@app.route('/process-files', methods=['POST'])
def upload_file():
    global progress
    progress = 0

    # Check if folder and Excel file parts are in the request
    if 'folderUpload' not in request.files or 'excelFile' not in request.files:
        return jsonify(message='No folder or Excel file part'), 400

    excel_file = request.files['excelFile']
    if excel_file.filename == '':
        return jsonify(message='No selected Excel file'), 400

    # Save Excel file
    if excel_file and allowed_file(excel_file.filename):
        excel_filename = secure_filename(excel_file.filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        excel_file.save(excel_path)
    else:
        return jsonify(message='Invalid Excel file format'), 400

    # Create folder for uploaded files
    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], 'uploaded_folder')
    os.makedirs(folder_path, exist_ok=True)

    folder_files = request.files.getlist('folderUpload')
    if not folder_files:
        return jsonify(message='No files in the folder'), 400

    # Save each file in the folder
    for folder_file in folder_files:
        if folder_file and allowed_file(folder_file.filename) and 'passport' in folder_file.filename.lower():
            folder_filename = secure_filename(os.path.basename(folder_file.filename))
            file_path = os.path.join(folder_path, folder_filename)
            folder_file.save(file_path)

    # Excel output path
    new_excel_filename = "new.xlsx"
    excel_output_path = os.path.join(folder_path, new_excel_filename)

    # Check if Excel file exists, otherwise create a new workbook
    if not os.path.exists(excel_output_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Passport Data'
        headers = [
            "Role in EWC", "Game", "Purpose of Visit", "Given Name", "Last Name",
            "Gender", "Date of Birth", "Place of Birth", "Nationality", "Arrival Date", "Departure Date",
            "Departure City", "Departure Airport", "Duration of Stay", "Entry Type [single, multiple]",
            "Address of Residence in KSA", "Passport Number", "Passport Issuing Date", "Passport Expiry Date",
            "Passport Issuance City", "Passport Type", "Phone Number", "Email", "Scan Accuracy", "File Name", "File Path",
            "MRZ1", "MRZ2"
        ]
        sheet.append(headers)

        # Styling the headers
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index)
            cell.value = header
            cell.fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add border to header cell
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
            cell.border = border

        workbook.save(excel_output_path)
    else:
        # Load existing workbook
        workbook = load_workbook(excel_output_path)
        sheet = workbook.active

    # Process files using the new Excel file
    output_excel_path = process_files(folder_path, excel_output_path)
    output_filename = os.path.basename(output_excel_path)

    # Copy columns from source to target based on passport number
    source_file_path = excel_path  # Use uploaded Excel file as source
    target_file_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
    copy_columns_based_on_passport(source_file_path, target_file_path)

    return jsonify(message='Files processed successfully', excel_path=url_for('download_file', filename=output_filename))

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(app.config['PROCESSED_FOLDER'], filename), as_attachment=True)

@app.route('/progress')
def progress_stream():
    def generate():
        global progress
        while progress < 100:
            time.sleep(1)
            yield f"data:{progress}\n\n"
    return Response(generate(), mimetype='text/event-stream')

def process_files(folder_path, excel_path):
    global progress
    total_files = len([f for f in os.listdir(folder_path) if 'passport' in f.lower()])
    
    # Load the Excel workbook and select the active sheet
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    next_row = sheet.max_row + 1

    # Process each file in the folder
    for index, file_name in enumerate(os.listdir(folder_path)):
        if 'passport' in file_name.lower():
            file_path = os.path.join(folder_path, file_name)
            process_passport_file(file_path, sheet, next_row)
            next_row += 1
            progress = int(((index + 1) / total_files) * 100)

    # Highlight cells containing "CHECK"
    highlight_cells_with_check(sheet)
    excel_output_path = os.path.join(app.config['PROCESSED_FOLDER'], 'processed_output.xlsx')
    workbook.save(excel_output_path)

    progress = 100
    print(f"All passport information extracted and saved to {excel_output_path}")
    return excel_output_path

def process_passport_file(file_path, sheet, next_row):
    # Check if file is a valid image file
    valid_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']
    if not any(file_path.lower().endswith(ext) for ext in valid_extensions):
        print(f"Skipping file: {file_path}. File type not allowed.")
        return

    # Extract file name from file path
    file_name = os.path.basename(file_path)

    # Load and parse the file
    input_doc = mindee_client.source_from_path(file_path)
    result = mindee_client.parse(product.PassportV1, input_doc)

    # Extract relevant information
    passport_info = result.document.inference.prediction
    accuracy = (passport_info.mrz1.confidence + passport_info.mrz2.confidence) / 2

    # Check MRZ length for TD3 format
    if len(str(passport_info.mrz1)) != 44 or len(str(passport_info.mrz2)) != 44:
        data = {
            "Role in EWC": "",
            "Game": "",
            "Purpose of Visit": "",
            "Given Name": "CHECK",
            "Last Name": "CHECK",
            "Gender": "CHECK",
            "Date of Birth": "CHECK",
            "Place of Birth": "CHECK",
            "Nationality": "CHECK",
            "Arrival Date": "",
            "Departure Date": "",
            "Departure City": "",
            "Departure Airport": "",
            "Duration of Stay": "",
            "Entry Type [single, multiple]": "",
            "Address of Residence in KSA": "",
            "Passport Number": "CHECK",
            "Passport Issuing Date": "",
            "Passport Expiry Date": "CHECK",
            "Passport Issuance City": "",
            "Passport Type": "Regular",
            "Phone Number": "",
            "Email": "",
            "Scan Accuracy": accuracy * 100,
            "File Name": file_name,
            "File Path": file_path,
            "MRZ1": str(passport_info.mrz1),
            "MRZ2": str(passport_info.mrz2)
        }
        append_to_excel(sheet, data, next_row)
        return

    # Create TD3CodeChecker instance with MRZ data
    td3_check = TD3CodeChecker(str(passport_info.mrz1) + "\n" + str(passport_info.mrz2))
    mrz_fields = td3_check.fields()

    # Extract country code and map it to full country name
    country_code = str(mrz_fields.country)
    country_name = country_codes.get(country_code, "") if country_code else country_code

    data = {
        "Role in EWC": "",
        "Game": "",
        "Purpose of Visit": "",
        "Given Name": mrz_fields.name if mrz_fields.name != 'None' else "CHECK",
        "Last Name": mrz_fields.surname if mrz_fields.surname != 'None' else "CHECK",
        "Gender": mrz_fields.sex,
        "Date of Birth": convert_yymmdd_to_ddmmyyyy(str(mrz_fields.birth_date)),
        "Place of Birth": country_name,
        "Nationality": get_country(mrz_fields.nationality),
        "Arrival Date": "",
        "Departure Date": "",
        "Departure City": "",
        "Departure Airport": "",
        "Duration of Stay": "",
        "Entry Type [single, multiple]": "",
        "Address of Residence in KSA": "",
        "Passport Number": mrz_fields.document_number,
        "Passport Issuing Date": "",
        "Passport Expiry Date": convert_yymmdd_to_ddmmyyyy(str(mrz_fields.expiry_date)),
        "Passport Issuance City": "",
        "Passport Type": "Regular",
        "Phone Number": "",
        "Email": "",
        "Scan Accuracy": accuracy * 100,
        "File Name": file_name,
        "File Path": file_path,
        "MRZ1": str(passport_info.mrz1),
        "MRZ2": str(passport_info.mrz2)
    }
    append_to_excel(sheet, data, next_row)

def append_to_excel(sheet, data, next_row):
    # Convert data keys to a list
    data_keys_list = list(data.keys())

    # Write data to the sheet
    for col_index, key in enumerate(data_keys_list, start=1):
        cell = sheet.cell(row=next_row, column=col_index)
        cell.value = data[key]
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add hyperlink to the 'File Name' column
    file_index = data_keys_list.index("File Name") + 1
    file_cell = sheet.cell(row=next_row, column=file_index)
    file_hyperlink = Hyperlink(ref=f"{data['File Path']}", target=f"{data['File Path']}")
    file_cell.value = data['File Name']
    file_cell.style = "Hyperlink"
    file_cell.hyperlink = file_hyperlink

    # Apply conditional formatting to Passport Expiry Date column
    expiry_date_index = data_keys_list.index("Passport Expiry Date") + 1
    expiry_cell = sheet.cell(row=next_row, column=expiry_date_index)

    expiry_date_str = data["Passport Expiry Date"]
    if expiry_date_str != "CHECK":
        try:
            expiry_date = datetime.strptime(expiry_date_str, "%d-%m-%Y").date()
            today = date.today()
            six_months_from_today = today + timedelta(days=180)
            if expiry_date <= six_months_from_today:
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                expiry_cell.fill = red_fill
        except ValueError:
            print(f"Error parsing Passport Expiry Date: {expiry_date_str}. Invalid date format.")

def highlight_cells_with_check(sheet):
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "CHECK" in cell.value.upper():
                cell.fill = red_fill

def convert_yymmdd_to_ddmmyyyy(date_str):
    try:
        date_obj = datetime.strptime(date_str, "%y%m%d")
        formatted_date = date_obj.strftime("%d-%m-%Y")
        return formatted_date
    except ValueError:
        print(f"Error parsing date: {date_str}. Invalid format.")
        return None

def copy_columns_based_on_passport(source_file, target_file):
    # Load source workbook and target workbook
    source_wb = load_workbook(source_file)
    target_wb = load_workbook(target_file)

    # Assuming both files have a sheet named 'Passport Data' (change as needed)
    source_sheet = source_wb['Passport Data']
    target_sheet = target_wb['Passport Data']

    # Columns to copy
    columns_to_copy = [
        "Role in EWC", "Game", "Purpose of Visit", "Arrival Date", "Departure Date",
        "Departure City", "Departure Airport", "Duration of Stay", "Entry Type [single, multiple]",
        "Address of Residence in KSA", "Passport Issuance City", "Passport Type", "Phone Number", "Email"
    ]

    # Find index of Passport Number column in both sheets
    passport_col_index_source = 0
    passport_col_index_target = 0
    for col_idx in range(1, source_sheet.max_column + 1):
        if source_sheet.cell(row=1, column=col_idx).value == "Passport Number":
            passport_col_index_source = col_idx
            break

    for col_idx in range(1, target_sheet.max_column + 1):
        if target_sheet.cell(row=1, column=col_idx).value == "Passport Number":
            passport_col_index_target = col_idx
            break

    if passport_col_index_source == 0 or passport_col_index_target == 0:
        print("Passport Number column not found in one of the sheets.")
        return

    # Iterate through rows in source sheet and copy data to target sheet based on Passport Number match
    for source_row_idx in range(2, source_sheet.max_row + 1):  # Start from row 2 to skip header
        source_passport_number = source_sheet.cell(row=source_row_idx, column=passport_col_index_source).value
        if not source_passport_number:
            continue

        for target_row_idx in range(2, target_sheet.max_row + 1):  # Start from row 2 to skip header
            target_passport_number = target_sheet.cell(row=target_row_idx, column=passport_col_index_target).value
            if not target_passport_number:
                continue

            if source_passport_number == target_passport_number:
                # Copy data from source to target for each column
                for col_name in columns_to_copy:
                    source_col_idx = 0
                    target_col_idx = 0
                    for col_idx in range(1, source_sheet.max_column + 1):
                        if source_sheet.cell(row=1, column=col_idx).value == col_name:
                            source_col_idx = col_idx
                            break

                    for col_idx in range(1, target_sheet.max_column + 1):
                        if target_sheet.cell(row=1, column=col_idx).value == col_name:
                            target_col_idx = col_idx
                            break

                    if source_col_idx == 0 or target_col_idx == 0:
                        print(f"Column '{col_name}' not found in one of the sheets.")
                        continue

                    # Copy value from source to target
                    source_cell_value = source_sheet.cell(row=source_row_idx, column=source_col_idx).value
                    target_sheet.cell(row=target_row_idx, column=target_col_idx).value = source_cell_value

    # Save the target workbook
    target_wb.save(target_file)
    print(f"Columns copied based on Passport Number from {os.path.basename(source_file)} to {os.path.basename(target_file)}.")

if __name__ == "__main__":
    app.run(debug=True)
